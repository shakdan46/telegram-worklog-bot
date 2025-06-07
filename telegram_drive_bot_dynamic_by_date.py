import logging
import io
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ApplicationBuilder, CommandHandler, ContextTypes, MessageHandler, filters, CallbackQueryHandler, ConversationHandler
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload
from datetime import datetime
import openpyxl
import os

# === שלבים ===
SELECTING_DATE, SELECTING_WORKERS = range(2)

# === קביעות ===
SCOPES = ['https://www.googleapis.com/auth/drive']
FILE_ID = '1UT4zWtuny8ES2z3dF2UTudK86HF1Mm5B'
MONTH_MAP = {
    '01': 'ינואר', '02': 'פברואר', '03': 'מרץ', '04': 'אפריל',
    '05': 'מאי', '06': 'יוני', '07': 'יולי', '08': 'אוגוסט',
    '09': 'ספטמבר', '10': 'אוקטובר', '11': 'נובמבר', '12': 'דצמבר'
}

# התחברות ל-Google Drive
CREDENTIALS_JSON = os.environ.get('GOOGLE_APPLICATION_CREDENTIALS_JSON')
credentials = service_account.Credentials.from_service_account_info(eval(CREDENTIALS_JSON), scopes=SCOPES)
drive_service = build('drive', 'v3', credentials=credentials)

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("שלום! אנא הזן תאריך בפורמט: YYYY-MM-DD")
    return SELECTING_DATE

async def receive_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    date_text = update.message.text.strip()
    try:
        date_obj = datetime.strptime(date_text, '%Y-%m-%d')
        hebrew_month = MONTH_MAP[date_obj.strftime('%m')]
        context.user_data['selected_date'] = date_obj.strftime('%d/%m/%Y')
        context.user_data['month_name'] = hebrew_month

        request = drive_service.files().get_media(fileId=FILE_ID)
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            status, done = downloader.next_chunk()
        fh.seek(0)

        wb = openpyxl.load_workbook(fh)
        sheet = wb[hebrew_month]

        workers = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_date = row[0].strftime('%d/%m/%Y') if isinstance(row[0], datetime) else row[0]
            if row_date == context.user_data['selected_date'] and row[1]:
                workers.add(str(row[1]).strip())

        if not workers:
            await update.message.reply_text("⚠️ לא נמצאו פועלים בתאריך הזה.")
            return ConversationHandler.END

        context.user_data['workers'] = list(workers)
        context.user_data['selected'] = []

        buttons = [[InlineKeyboardButton(name, callback_data=name)] for name in workers]
        buttons.append([InlineKeyboardButton("סיום ✅", callback_data="done")])
        await update.message.reply_text("בחר את העובדים:", reply_markup=InlineKeyboardMarkup(buttons))
        return SELECTING_WORKERS

    except Exception as e:
        print(e)
        await update.message.reply_text("❌ שגיאה: ודא שכתבת תאריך חוקי בפורמט YYYY-MM-DD")
        return SELECTING_DATE

async def handle_worker_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    choice = query.data

    if choice == "done":
        selected = context.user_data.get("selected", [])
        if not selected:
            await query.edit_message_text(text="לא נבחרו עובדים.")
            return ConversationHandler.END

        request = drive_service.files().get_media(fileId=FILE_ID)
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            status, done = downloader.next_chunk()
        fh.seek(0)

        wb = openpyxl.load_workbook(fh)
        sheet = wb[context.user_data['month_name']]

        for row in sheet.iter_rows(min_row=2):
            date_cell = row[0]
            name_cell = row[1]
            status_cell = row[2]  # עמודת "הגיע לעבודה?"
            date_val = date_cell.value.strftime('%d/%m/%Y') if isinstance(date_cell.value, datetime) else date_cell.value
            name_val = str(name_cell.value).strip() if name_cell.value else ""
            if date_val == context.user_data['selected_date'] and name_val in selected:
                status_cell.value = True

        out_stream = io.BytesIO()
        wb.save(out_stream)
        out_stream.seek(0)

        media_body = MediaIoBaseUpload(out_stream, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        drive_service.files().update(fileId=FILE_ID, media_body=media_body).execute()

        await query.edit_message_text(text="העובדים שנבחרו ✅:\n" + "\n".join(selected))
        return ConversationHandler.END

    if choice not in context.user_data["selected"]:
        context.user_data["selected"].append(choice)

    await query.answer(text=f"{choice} נוסף ✅", show_alert=False)
    return SELECTING_WORKERS

if __name__ == '__main__':
    from dotenv import load_dotenv
    load_dotenv()
    TOKEN = os.environ.get("TELEGRAM_TOKEN")
    app = ApplicationBuilder().token(TOKEN).build()
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler('start', start)],
        states={
            SELECTING_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_date)],
            SELECTING_WORKERS: [CallbackQueryHandler(handle_worker_selection)],
        },
        fallbacks=[]
    )
    app.add_handler(conv_handler)
    app.run_polling()

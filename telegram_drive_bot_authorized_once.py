# telegram_drive_bot_authorized_once.py
import os
import io
import json
import logging
from datetime import datetime, timedelta
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    ApplicationBuilder, CommandHandler, CallbackQueryHandler,
    ContextTypes, ConversationHandler, MessageHandler, filters
)
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload
from openpyxl import load_workbook
from dotenv import load_dotenv

# === אתחול ===
load_dotenv()
logging.basicConfig(level=logging.INFO)

# === קבועים ===
SCOPES = ['https://www.googleapis.com/auth/drive']
FILE_ID = os.environ.get("EXCEL_FILE_ID")
CREDENTIALS_JSON = json.loads(os.environ.get("GOOGLE_APPLICATION_CREDENTIALS_JSON"))
TOKEN = os.environ.get("TELEGRAM_TOKEN")
PASSWORD = '204560916'
AUTHORIZED_USERS = set()

(ASK_PASSWORD, SELECTING_DATE, SELECTING_WORKERS, CONFIRM_SELECTION,
 ADD_NAME, ADD_SALARY, ADD_START_DATE) = range(7)

MONTH_MAP = {
    1: "ינואר", 2: "פברואר", 3: "מרץ", 4: "אפריל", 5: "מאי", 6: "יוני",
    7: "יולי", 8: "אוגוסט", 9: "ספטמבר", 10: "אוקטובר", 11: "נובמבר", 12: "דצמבר"
}

credentials = service_account.Credentials.from_service_account_info(
    CREDENTIALS_JSON, scopes=SCOPES)
drive_service = build('drive', 'v3', credentials=credentials)

# === הורדת והעלאת קובץ ===
def download_excel():
    request = drive_service.files().get_media(fileId=FILE_ID)
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)
    while not downloader.next_chunk()[1]:
        pass
    fh.seek(0)
    return load_workbook(fh)

def upload_excel(wb):
    out_stream = io.BytesIO()
    wb.save(out_stream)
    out_stream.seek(0)
    media_body = MediaIoBaseUpload(out_stream,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    drive_service.files().update(fileId=FILE_ID, media_body=media_body).execute()

# === התחלת בוט ===
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if user_id in AUTHORIZED_USERS:
        await update.message.reply_text("הזן תאריך בפורמט YYYY-MM-DD:")
        return SELECTING_DATE
    else:
        await update.message.reply_text("🔐 אנא הזן סיסמה:")
        return ASK_PASSWORD

async def check_password(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text.strip() == PASSWORD:
        AUTHORIZED_USERS.add(update.effective_user.id)
        await update.message.reply_text("✅ הצלחת להזדהות. הזן תאריך (YYYY-MM-DD):")
        return SELECTING_DATE
    else:
        await update.message.reply_text("❌ סיסמה שגויה.")
        return ASK_PASSWORD

# === תאריך ועובדים ===
async def receive_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        date_obj = datetime.strptime(update.message.text.strip(), '%Y-%m-%d')
        month_name = MONTH_MAP[date_obj.month]
        selected_date = date_obj.strftime('%d/%m/%Y')

        context.user_data['selected_date'] = selected_date
        context.user_data['month_name'] = month_name

        wb = download_excel()
        sheet = wb[month_name]

        workers = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                row_date = row[0].strftime('%d/%m/%Y') if isinstance(row[0], datetime) else row[0]
                if row_date == selected_date:
                    workers.append(row[1])

        if not workers:
            await update.message.reply_text("⚠️ לא נמצאו פועלים בתאריך הזה.")
            return ConversationHandler.END

        context.user_data["all_workers"] = list(set(workers))
        context.user_data["selected"] = []
        return await show_worker_selection(update, context)

    except Exception as e:
        await update.message.reply_text("❌ תאריך שגוי. נסה שוב:")
        return SELECTING_DATE

async def show_worker_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    selected = context.user_data.get("selected", [])
    available = [w for w in context.user_data["all_workers"] if w not in selected]
    buttons = [[InlineKeyboardButton(w, callback_data=w)] for w in available]
    buttons.append([InlineKeyboardButton("סיום ✅", callback_data="done")])

    msg = "בחר עובדים:"
    if update.callback_query:
        await update.callback_query.edit_message_text(msg, reply_markup=InlineKeyboardMarkup(buttons))
    else:
        await update.message.reply_text(msg, reply_markup=InlineKeyboardMarkup(buttons))
    return SELECTING_WORKERS

async def handle_worker_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    choice = query.data
    await query.answer()

    if choice == "done":
        if not context.user_data["selected"]:
            await query.edit_message_text("❗ לא נבחרו פועלים.")
            return ConversationHandler.END
        return await show_confirmation_menu(query, context)

    if choice not in context.user_data["selected"]:
        context.user_data["selected"].append(choice)

    return await show_worker_selection(update, context)

async def show_confirmation_menu(query, context):
    selected = context.user_data["selected"]
    text = "✅ העובדים שנבחרו:\n" + "\n".join(selected)
    buttons = [
        [InlineKeyboardButton("אישור ✅", callback_data="confirm_final")],
        [InlineKeyboardButton("הסר שם ❌", callback_data="remove_worker")],
        [InlineKeyboardButton("➕ הוסף פועל נוסף", callback_data="add_worker")]
    ]
    await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(buttons))
    return CONFIRM_SELECTION

async def confirmation_router(update: Update, context: ContextTypes.DEFAULT_TYPE):
    data = update.callback_query.data
    if data == "confirm_final":
        return await confirm_selection(update, context)
    elif data == "remove_worker":
        return await remove_worker(update, context)
    elif data.startswith("remove_"):
        return await remove_worker_choice(update, context)
    elif data == "add_worker":
        return await show_worker_selection(update, context)

async def confirm_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    selected = context.user_data["selected"]
    wb = download_excel()
    sheet = wb[context.user_data["month_name"]]
    target_date = context.user_data["selected_date"]

    for row in sheet.iter_rows(min_row=2):
        date_cell = row[0]
        name_cell = row[1]
        if date_cell.value and name_cell.value:
            row_date = date_cell.value.strftime('%d/%m/%Y') if isinstance(date_cell.value, datetime) else date_cell.value
            if row_date == target_date and str(name_cell.value).strip() in selected:
                row[2].value = True  # עמודת "הגיע לעבודה?"

    upload_excel(wb)
    await query.edit_message_text("🎉 הנתונים נשמרו:\n" + "\n".join(selected))
    return ConversationHandler.END

async def remove_worker(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    selected = context.user_data["selected"]
    buttons = [[InlineKeyboardButton(f"{w} ❌", callback_data=f"remove_{w}")] for w in selected]
    await query.edit_message_text("בחר שם להסרה:", reply_markup=InlineKeyboardMarkup(buttons))
    return CONFIRM_SELECTION

async def remove_worker_choice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    name = update.callback_query.data.replace("remove_", "")
    context.user_data["selected"].remove(name)
    return await show_confirmation_menu(update.callback_query, context)

# === פקודת /addworker להוספת פועל חדש ===
async def start_add_worker(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("מה שם הפועל החדש?")
    return ADD_NAME

async def add_worker_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["new_worker_name"] = update.message.text.strip()
    await update.message.reply_text("מה השכר היומי שלו?")
    return ADD_SALARY

async def add_worker_salary(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        context.user_data["new_worker_salary"] = float(update.message.text.strip())
        await update.message.reply_text("מה תאריך התחלה שלו? (dd/mm/yyyy)")
        return ADD_START_DATE
    except:
        await update.message.reply_text("❗ סכום שכר לא תקין.")
        return ADD_SALARY

async def add_worker_start_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        start_date = datetime.strptime(update.message.text.strip(), '%d/%m/%Y')
    except:
        await update.message.reply_text("❗ תאריך שגוי.")
        return ADD_START_DATE

    name = context.user_data["new_worker_name"]
    salary = context.user_data["new_worker_salary"]
    wb = download_excel()

    current = start_date
    while current.year == 2025:
        sheet_name = MONTH_MAP[current.month]
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            row = sheet.max_row + 1
            sheet.cell(row=row, column=1).value = current.strftime('%d/%m/%Y')
            sheet.cell(row=row, column=2).value = name
            sheet.cell(row=row, column=5).value = salary
        current += timedelta(days=1)

    if "יומית פועלים" in wb.sheetnames:
        ws = wb["יומית פועלים"]
        ws.append([name, salary])

    upload_excel(wb)
    await update.message.reply_text(f"✅ הפועל {name} נוסף עם שכר ₪{salary}.")
    return ConversationHandler.END

# === main ===
def main():
    app = ApplicationBuilder().token(TOKEN).build()

    app.add_handler(ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            ASK_PASSWORD: [MessageHandler(filters.TEXT & ~filters.COMMAND, check_password)],
            SELECTING_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_date)],
            SELECTING_WORKERS: [CallbackQueryHandler(handle_worker_selection)],
            CONFIRM_SELECTION: [CallbackQueryHandler(confirmation_router)],
        },
        fallbacks=[]
    ))

    app.add_handler(ConversationHandler(
        entry_points=[CommandHandler("addworker", start_add_worker)],
        states={
            ADD_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, add_worker_name)],
            ADD_SALARY: [MessageHandler(filters.TEXT & ~filters.COMMAND, add_worker_salary)],
            ADD_START_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, add_worker_start_date)],
        },
        fallbacks=[]
    ))

    app.run_polling()

if __name__ == '__main__':
    main()
